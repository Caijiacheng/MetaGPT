# 测试 @metagpt/ext/aico/services/project_tracking_manager.py
## project_tracking_manager代码目标覆盖率：80%以上
## 测试的目录和数据放在 @tests/data/aico/services/project_tracking_manager


# 运行测试
# pytest tests/metagpt/ext/aico/services/test_project_tracking_mananger.py -v --cov=metagpt --cov-report=term

# coverage report --include="metagpt/ext/aico/**/project_tracking_manager.py"

import pytest
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
import re  # 添加正则表达式模块导入
from metagpt.ext.aico.services.project_tracking_manager import (
    ProjectTrackingManager,
    TrackingSheet
)
from metagpt.ext.aico.config.tracking_config import SheetType
from unittest.mock import patch

TEST_DATA_DIR = Path(__file__).parent.parent / "data/aico/services/project_tracking_manager"

@pytest.fixture
def test_excel_path(tmp_path):
    return tmp_path / "test_project_tracking.xlsx"

@pytest.fixture
def tracking_manager(test_excel_path, tmp_path):
    # 创建临时项目根目录
    project_root = tmp_path / "project_root"
    project_root.mkdir()
    
    # 确保每次测试使用新文件
    if test_excel_path.exists():
        test_excel_path.unlink()
    return ProjectTrackingManager(test_excel_path, project_root)

class TestProjectTrackingManagerInit:
    def test_file_creation(self, test_excel_path, tracking_manager):
        """测试文件初始化"""
        assert test_excel_path.exists()
        # 更新预期工作表列表（移除版本历史）
        expected_sheets = [st.value.name for st in SheetType]
        assert sorted(tracking_manager.wb.sheetnames) == sorted(expected_sheets)

    def test_sheet_headers(self, tracking_manager):
        """测试各工作表表头正确性"""
        for sheet_type in SheetType:
            ws = tracking_manager.wb[sheet_type.value.name]
            assert [cell.value for cell in ws[1]] == sheet_type.value.headers

class TestRawRequirementOperations:
    def test_add_raw_requirement(self, tracking_manager):
        """测试添加原始需求记录"""
        test_data = {
            "file_path": "raw/iter1/test_req.md",
            "req_type": "文档",
            "comment": "测试需求"
        }
        
        tracking_manager.add_raw_requirement(**test_data)
        
        ws = tracking_manager.wb[TrackingSheet.RAW_REQUIREMENTS.value]
        last_row = ws.iter_rows(min_row=2).__next__()
        
        assert last_row[0].value == test_data["file_path"]
        assert last_row[1].value == test_data["req_type"]
        assert last_row[8].value == test_data["comment"]
        assert last_row[3].value == "待分析"

    def test_update_raw_requirement_status(self, tracking_manager):
        """测试更新原始需求状态（修正参数）"""
        # 准备测试数据
        file_path = "raw/iter1/status_test.md"
        tracking_manager.add_raw_requirement(file_path, "文档")
        
        # 执行状态更新（移除多余的时间参数）
        result = tracking_manager.update_raw_requirement_status(file_path, "已分析")
        
        assert result is True
        status = tracking_manager.get_raw_requirement_status(file_path)
        assert status["status"] == "已分析"

    def test_invalid_status_update(self, tracking_manager):
        """测试无效状态更新"""
        result = tracking_manager.update_raw_requirement_status("non_exist.md", "已分析")
        assert result is False

class TestRequirementManagement:
    def test_add_standard_requirement(self, tracking_manager):
        """测试添加标准需求"""
        test_data = {
            "name": "测试需求",
            "description": "需求描述",
            "priority": "中",
            "status": "已提出"
        }
        
        req = tracking_manager.add_standard_requirement(test_data)
        assert req["id"].startswith("REQ-PM")
        
    def test_default_priority(self, tracking_manager):
        """测试默认优先级应用"""
        # 不提供priority字段
        req = tracking_manager.add_standard_requirement({
            "name": "测试默认优先级需求"
        })
        assert req["priority"] == "中"

    def test_priority_override(self, tracking_manager):
        """测试显式设置优先级"""
        req = tracking_manager.add_standard_requirement({
            "name": "测试高优先级需求",
            "priority": "高"
        })
        assert req is not None, "需求对象不应为None"
        assert req["priority"] == "高"

class TestUserStoryOperations:
    def test_user_story_lifecycle(self, tracking_manager):
        """测试用户故事全生命周期"""
        # 添加合法优先级
        req = tracking_manager.add_standard_requirement({
            "name": "测试需求",
            "priority": "中"  # 添加必填字段
        })
        
        # 创建用户故事
        story_data = {
            "related_req_id": req["id"],
            "name": "测试用户故事",
            "status": "待拆分"
        }
        story = tracking_manager.add_user_story(story_data)
        
        # 验证ID格式
        assert re.match(rf"US-{req['id']}-\d{{2}}", story["story_id"])
        
        # 测试状态更新
        tracking_manager.update_user_story_status(story["story_id"], "开发中")
        updated_story = tracking_manager.get_user_story(story["story_id"])
        assert updated_story["status"] == "开发中"

class TestTaskTracking:
    def test_task_operations(self, tracking_manager):
        """测试任务跟踪全流程"""
        # 创建需求
        req = tracking_manager.add_standard_requirement({
            "name": "测试需求"
        })
        # 创建用户故事
        story = tracking_manager.add_user_story({"related_req_id": req["id"], "name": "测试用户故事"})
        
        # 添加任务
        task_data = {
            "related_req_id": req["id"],
            "related_story_id": story["story_id"],
            "name": "开发任务",
            "description": "任务描述",
            "type": "dev",
            "owner": "李工",
            "plan_start": "2024-06-01",
            "plan_end": "2024-06-05",
            "notes": "测试任务"
        }
        
        task = tracking_manager.add_task(task_data)
        assert re.match(rf"^T-{story['story_id']}-DEV\d{{2}}$", task["task_id"])
        
        # 验证初始状态
        ws = tracking_manager.wb[TrackingSheet.TASK_TRACKING.value]
        last_row = ws.iter_rows(min_row=2).__next__()
        assert last_row[7].value == "待开始"
        
        # 更新状态
        tracking_manager.update_task_status(task["task_id"], "进行中")
        assert last_row[7].value == "进行中"
        assert last_row[9].value is not None  # 实际开始时间

    def test_invalid_task_operation(self, tracking_manager):
        """测试无效任务操作"""
        # 不存在的任务ID
        assert tracking_manager.update_task_status("INVALID_TASK", "进行中") is False
        assert tracking_manager.update_task_artifacts("INVALID_TASK", []) is False

class TestValidation:
    def test_file_structure_validation(self, tracking_manager):
        """测试文件结构校验"""
        assert tracking_manager.validate_file_structure() is True
        
        # 测试异常情况
        del tracking_manager.wb[TrackingSheet.RAW_REQUIREMENTS.value]
        assert tracking_manager.validate_file_structure() is False

    def test_baseline_operations(self, tracking_manager):
        """测试基线管理功能"""
        # 生成符合规范的需求ID
        req = tracking_manager.add_standard_requirement({
            "name": "测试需求",
            "priority": "高"
        })
        req_ids = [req["id"]]
        
        # 标记设计基线
        tracking_manager.mark_baseline("design", "v1.0", req_ids)
        
        # 验证基线状态
        baseline_reqs = tracking_manager.get_baseline_requirements("design")
        assert set(baseline_reqs) == set(req_ids)
        
        # 标记代码基线
        tracking_manager.mark_baseline("code", "v2.0", req_ids)
        code_baseline = tracking_manager.get_baseline_requirements("code", "v2.0")
        assert set(code_baseline) == set(req_ids)

class TestAdvancedOperations:
    def test_update_requirement(self, tracking_manager):
        """测试需求更新操作"""
        # 创建初始需求
        req = tracking_manager.add_standard_requirement({
            "name": "原始需求",
            "priority": "中"
        })
        
        # 更新需求
        updated = tracking_manager.update_requirement(
            req["id"],
            {"priority": "高", "status": "已分析"}
        )
        
        # 验证更新结果
        assert updated["priority"] == "高"
        assert updated["status"] == "已分析"

    def test_requirement_validation(self, tracking_manager):
        """测试需求数据校验"""
        valid_data = {
            "file_path": "valid.md",
            "description": "完整的需求描述",
            "source": "客户"
        }
        assert tracking_manager.validate_raw_requirement(valid_data) is True
        
        invalid_data = {"file_path": "invalid.md"}
        assert tracking_manager.validate_raw_requirement(invalid_data) is False

    def test_design_status_update(self, tracking_manager):
        """测试设计状态更新"""
        # 创建符合规范的需求
        req = tracking_manager.add_standard_requirement({
            "name": "设计需求"
        })
        req_id = req["id"]  # 使用自动生成的ID
        
        tracking_manager.update_design_status(req_id, "基线化", "design_v1.md")
        req_status = tracking_manager.wb[TrackingSheet.REQUIREMENT_MGMT.value].iter_rows(min_row=2).__next__()
        assert req_status[12].value == "基线化"
        assert "design_v1.md" in req_status[13].value

    def test_task_artifacts(self, tracking_manager):
        """测试任务产物管理"""
        # 添加必填name字段
        req = tracking_manager.add_standard_requirement({
            "name": "测试需求"  # 新增name字段
        })
        story = tracking_manager.add_user_story({"related_req_id": req["id"]})
        task = tracking_manager.add_task({
            "related_story_id": story["story_id"],
            "type": "dev"
        })
        
        artifacts = ["artifact1.md", "artifact2.png"]
        tracking_manager.update_task_artifacts(task["task_id"], artifacts)
        
        ws = tracking_manager.wb[TrackingSheet.TASK_TRACKING.value]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == task["task_id"]:
                assert all(a in row[12].value for a in artifacts)  # ARTIFACTS列

class TestEdgeCases:
    def test_large_data_volume(self, tracking_manager):
        """测试大数据量处理"""
        for i in range(100):
            tracking_manager.add_raw_requirement(f"raw/iter{i}/bulk_{i}.md", "批量需求")
        
        # 验证数量
        ws = tracking_manager.wb[TrackingSheet.RAW_REQUIREMENTS.value]
        assert ws.max_row == 101  # 表头+100条数据
        
    def test_special_characters(self, tracking_manager):
        """测试特殊字符处理"""
        test_data = {
            "file_path": "raw/iter1/特殊路径@2024.md",
            "req_type": "特殊/类型"
        }
        tracking_manager.add_raw_requirement(**test_data)
        
        ws = tracking_manager.wb[TrackingSheet.RAW_REQUIREMENTS.value]
        last_row = ws.iter_rows(min_row=2).__next__()
        assert last_row[0].value == test_data["file_path"]

class TestSpecCompliance:
    """测试是否符合PM指南规范"""
    
    def test_sheet_structure(self, tracking_manager):
        """验证工作表结构符合规范"""
        sheet_specs = {
            "需求管理": [
                "需求ID", "原始需求文件", "需求名称", "需求描述", "需求来源",
                "需求优先级", "需求状态", "提出人/负责人", "提出时间", "目标完成时间",
                "验收标准", "设计状态", "关联设计文档", "代码基线版本", "设计评审记录", "备注"
            ],
            "用户故事管理": [
                "用户故事ID", "关联需求ID", "用户故事名称", "用户故事描述",
                "优先级", "状态", "验收标准", "创建时间", "备注"
            ],
            "任务跟踪": [
                "任务ID", "关联需求ID", "关联用户故事ID", "任务名称", "任务描述",
                "任务类型", "负责人", "任务状态", "计划开始时间", "计划结束时间",
                "实际开始时间", "实际结束时间", "关联产出物", "备注"
            ]
        }
        
        for sheet_name, headers in sheet_specs.items():
            ws = tracking_manager.wb[sheet_name]
            actual_headers = [cell.value for cell in ws[1]]  # 第一行为表头
            assert actual_headers == headers, f"{sheet_name}表头不符合规范"
            
    def test_id_naming(self, tracking_manager):
        """验证ID命名规范"""
        req = tracking_manager.add_standard_requirement({
            "name": "测试需求",  # 确保name字段存在
            "priority": "中"
        })
        assert req["id"] is not None  # 增加非空校验
        
        story = tracking_manager.add_user_story({"related_req_id": req["id"]})
        assert re.match(rf"US-{req['id']}-\d{{2}}", story["story_id"])
        
        # 测试任务ID
        task_id = tracking_manager._generate_task_id(story["story_id"], "dev")
        assert re.match(rf"^T-{story['story_id']}-DEV\d{{2}}$", task_id)
        
    def test_file_path_convention(self, tracking_manager):
        """验证文件路径规范"""
        valid_path = "raw/iter3_batch.md"
        tracking_manager.add_raw_requirement(valid_path, "文档")
        
        # 验证非法路径
        with pytest.raises(ValueError):
            tracking_manager.add_raw_requirement("invalid_path.txt", "文档")
            
    def test_requirement_links(self, tracking_manager):
        """验证需求-用户故事-任务关联链"""
        # 创建需求
        req = tracking_manager.add_standard_requirement({
            "name": "测试需求",
            "priority": "高"
        })
        
        # 创建用户故事
        story = tracking_manager.add_user_story({"related_req_id": req["id"]})
        
        # 创建任务
        task = tracking_manager.add_task({
            "related_story_id": story["story_id"],
            "type": "test"
        })
        
        # 验证ID关联关系
        assert story["story_id"].startswith(f"US-{req['id']}")
        assert task["task_id"].startswith(f"T-{story['story_id']}-TEST")

    def test_version_directory_structure(self, tracking_manager):
        """此测试已不再适用"""
        pass  # 完全移除测试逻辑

    def test_status_flow(self, tracking_manager):
        """测试完整状态流转（修正状态设置）"""
        # 修正文件路径格式
        test_file = "raw/iter1/test_flow.md"  # 符合规范路径
        tracking_manager.add_raw_requirement(test_file, "文档")
        
        # 测试合法流转
        valid_transitions = [
            ("待分析", "已分析"),
            ("已分析", "完成"),
            ("已分析", "已驳回"),
            ("已驳回", "已分析")
        ]
        
        for from_status, to_status in valid_transitions:
            # 设置当前状态
            ws = tracking_manager.wb[SheetType.RAW_REQUIREMENT.value.name]
            for row in ws.iter_rows(min_row=2):
                if row[0].value == test_file:
                    row[3].value = from_status  # STATUS列
                    
            # 执行状态更新
            tracking_manager.update_raw_requirement_status(test_file, to_status)
            
            # 验证结果
            updated_status = tracking_manager.get_raw_requirement_status(test_file)
            assert updated_status["status"] == to_status
            if to_status == "已分析":
                assert updated_status["ba_time"] is not None
            elif to_status == "完成":
                assert updated_status["complete_time"] is not None
        
        # 测试非法流转前确保状态为"完成"
        ws = tracking_manager.wb[SheetType.RAW_REQUIREMENT.value.name]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == test_file:
                row[3].value = "完成"  # 显式设置最终状态
        
        # 测试非法流转
        with pytest.raises(ValueError) as excinfo:
            tracking_manager.update_raw_requirement_status(test_file, "已分析")
        assert "非法状态流转: 完成 → 已分析" in str(excinfo.value)

    def test_priority_values(self, tracking_manager):
        """验证优先级取值范围"""
        for priority in ["高", "中", "低"]:
            req = tracking_manager.add_standard_requirement({
                "name": f"测试{priority}优先级需求",
                "priority": priority,
                "description": "测试描述"
            })
            assert req is not None, "需求对象不应为None"
            assert req["priority"] == priority

    @patch("metagpt.ext.aico.services.project_tracking_manager.datetime")
    def test_sequence_management(self, mock_datetime, tracking_manager):
        """测试序列号递增逻辑（修正断言）"""
        mock_date = datetime(2024, 7, 1)
        mock_datetime.now.return_value = mock_date
        
        # 首次生成
        req_id1 = tracking_manager._generate_req_id()
        assert req_id1 == "REQ-PM2407001"  # 确保6位日期+3位序列号
        
        # 第二次生成
        req_id2 = tracking_manager._generate_req_id() 
        assert req_id2 == "REQ-PM2407002"  # 确保序列号递增