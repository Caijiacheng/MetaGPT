import unittest
from unittest.mock import patch, MagicMock
from pathlib import Path
import shutil
import os
from datetime import datetime
from openpyxl import load_workbook

from metagpt.ext.aico.roles.project_manager import AICOProjectManager
from metagpt.ext.aico.roles.business_analyst import AICOBusinessAnalyst
from metagpt.ext.aico.roles.enterprise_architect import AICOEnterpriseArchitect
from metagpt.ext.aico.services.doc_manager import DocManagerService, DocType
from metagpt.environment import Environment
from metagpt.config2 import config
from metagpt.ext.aico.config import config as aico_config

class TestAICOP0(unittest.IsolatedAsyncioTestCase):

    async def asyncSetUp(self):
        # 创建临时测试目录
        self.test_project_root = Path("test_project")
        self.test_project_root.mkdir(exist_ok=True)

        # 设置项目根目录
        aico_config.project_root = str(self.test_project_root)
        config.openai.api_key = "mock_key" # 设置假的api_key

        # 初始化 PM 角色
        self.pm = AICOProjectManager()
        self.pm.project_root = self.test_project_root  # 设置项目根目录
        self.pm.doc_manager = DocManagerService(self.test_project_root) # 传入项目根目录
        self.pm.rc.env = Environment()
        self.pm.rc.env.project_root = self.test_project_root

        # 初始化 BA 角色
        self.ba = AICOBusinessAnalyst()
        self.ba.project_root = self.test_project_root
        self.ba.doc_manager = DocManagerService(self.test_project_root)
        self.ba.rc.env = Environment()
        self.ba.rc.env.project_root = self.test_project_root

        # 初始化 EA 角色
        self.ea = AICOEnterpriseArchitect()
        self.ea.project_root = self.test_project_root
        self.ea.doc_manager = DocManagerService(self.test_project_root)
        self.ea.rc.env = Environment()
        self.ea.rc.env.project_root = self.test_project_root

    async def asyncTearDown(self):
        # 清理临时测试目录
        if self.test_project_root.exists():
            shutil.rmtree(self.test_project_root)

    @patch('metagpt.actions.action.Action.run')
    async def test_aico_p0_flow(self, mock_llm_run):
        # 模拟 LLM 响应
        mock_llm_run.side_effect = self.mock_llm_responses

        # 1. 项目初始化
        await self.pm._init_project()
        self.assertTrue(self.test_project_root.exists())

        # 使用 DocManagerService 获取跟踪表路径
        tracking_file_path = self.pm.doc_manager.get_doc_path(DocType.PROJECT_TRACKING)
        self.assertTrue(tracking_file_path.exists())

        version_file_path = self.test_project_root / "VERSION"
        self.assertTrue(version_file_path.exists())
        self.assertEqual(version_file_path.read_text().strip(), "0.1.0")

        # 2. 提交原始需求
        raw_req_file_path = self.pm.doc_manager.get_doc_path(DocType.REQUIREMENT_RAW) / "req-test-001.md"
        req_content = "# 测试需求\n\n用户希望能够搜索商品。"
        raw_req_file_path.parent.mkdir(parents=True, exist_ok=True)
        raw_req_file_path.write_text(req_content)

        req_input = {
            "file_path": str(raw_req_file_path.relative_to(self.test_project_root)),
            "description": "用户搜索商品",
            "source": "测试用例",
            "type": "business"
        }
        await self.pm._process_input_requirement(req_input)

        # 3. 业务需求分析
        await self.pm._parse_raw_requirements()
        await self.pm._process_requirements() # 包含了_parse_raw_requirements, _process_arch_analysis, _confirm_requirement_baseline

        # 断言跟踪表已更新
        tracking_file = self.test_project_root / "tracking/project_tracking.xlsx"
        self.assertTrue(tracking_file.exists())

        # 验证跟踪表中的数据
        wb = load_workbook(tracking_file)
        raw_req_sheet = wb["原始需求"]
        self.assertEqual(raw_req_sheet.cell(row=2, column=1).value, str(raw_req_file_path))  # 需求文件
        self.assertEqual(raw_req_sheet.cell(row=2, column=2).value, "business")  # 需求类型
        self.assertEqual(raw_req_sheet.cell(row=2, column=4).value, "parsed_by_ba")  # 当前状态

        # 4. 检查生成的文档
        # 检查版本是否递增
        version = self.pm.version_svc.current
        self.assertEqual(version, "0.2.0")

        # 使用 DocManagerService 获取需求分析报告路径
        biz_analysis_file = self.pm.doc_manager.get_doc_path(
            doc_type=DocType.REQUIREMENT_ANALYZED,
            version=version,
            req_id="REQ-20240101-001"  # 假设需求ID是这个
        )
        self.assertTrue(biz_analysis_file.exists())
        biz_analysis_content = biz_analysis_file.read_text()
        self.assertIn("需求概述", biz_analysis_content)  # 验证文档内容

        tech_analysis_file =  self.pm.doc_manager.get_doc_path(
            doc_type=DocType.TECH_REQUIREMENT,
            version=version,
            req_id="REQ-20240101-001"  # 假设需求ID是这个
        )
        self.assertTrue(tech_analysis_file.exists())

        biz_arch_file = self.pm.doc_manager.get_doc_path(
            doc_type=DocType.BUSINESS_ARCH,
            version=version
        )
        self.assertTrue(biz_arch_file.exists())

        tech_arch_file = self.pm.doc_manager.get_doc_path(
            doc_type=DocType.TECH_ARCH,
            version=version
        )
        self.assertTrue(tech_arch_file.exists())

        # 5. 需求基线确认
        # await self.pm._confirm_requirement_baseline() # 已在_process_requirements中执行
        # (在此添加对需求基线确认的断言，例如检查跟踪表中的状态)

    def mock_llm_responses(self, *args, **kwargs):
        """
        根据不同的Action，返回不同的模拟响应
        """
        action_class_name = args[0].__class__.__name__
        if action_class_name == "ParseBizRequirement":
            return MagicMock(
                content='{"business_architecture": "更新后的业务架构", "process_diagram": "Mermaid流程图", "user_stories": ["用户可以搜索商品"]}',
                instruct_content={
                    "requirement_id": "REQ-20240101-001",
                    "standard_requirements": [
                        {
                            "std_req_id": "SR-20240101001",
                            "description": "用户可以搜索商品",
                            "priority": "高"
                        }
                    ]
                }
            )
        elif action_class_name == "ParseTechRequirements":
            return MagicMock(
                content='{"requirements": ["技术需求1", "技术需求2"], "impact_analysis": "影响分析", "constraints": "约束"}',
                instruct_content={
                    "requirements": ["技术需求1", "技术需求2"],
                    "impact_analysis": "影响分析",
                    "constraints": "约束"
                },
                output_file = "path/to/matrix_file.xlsx"
            )
        elif action_class_name == "Update4ABusiness":
            return MagicMock(
                content='{"updated_architecture": "更新后的4A业务架构", "user_stories": {"SR-20240101001": [{"title": "搜索商品", "acceptance_criteria": "用户可以输入关键词搜索商品"}]}}',
                instruct_content={
                    "user_stories": {
                        "SR-20240101001": [
                            {
                                "story_id": "US-SR-20240101001-01",
                                "title": "搜索商品",
                                "status": "待评审",
                                "acceptance_criteria": "用户可以输入关键词搜索商品"
                            }
                        ]
                    },
                    "version": datetime.now().strftime("%Y%m%d%H%M")
                }
            )
        elif action_class_name == "Update4ATech":
            return MagicMock(
                content='{"architecture": {"application_architecture": {}, "data_architecture": {}, "technical_architecture": {}}, "components": [], "dependencies": [], "change_log": ["架构更新"]}',
                instruct_content={
                    "architecture": {
                        "application_architecture": {},
                        "data_architecture": {},
                        "technical_architecture": {}
                    },
                    "components": [],
                    "dependencies": [],
                    "change_log": ["架构更新"]
                }
            )
        elif action_class_name == "ReviewAllRequirements":
            return MagicMock(
                content='{"result": "评审通过"}',
                instruct_content={"result": "评审通过"}
            )

        return MagicMock(content="默认响应")

if __name__ == '__main__':
    unittest.main()
# ```

# **代码解释**:

# 1.  **`TestAICOP0` 类**:
#     *   继承自 `unittest.IsolatedAsyncioTestCase`，支持异步测试。
#     *   `asyncSetUp` 方法:
#         *   创建临时测试目录 `test_project`。
#         *   设置 `aico_config.project_root` 为测试目录。
#         *   初始化 `AICOProjectManager`、`AICOBusinessAnalyst` 和 `AICOEnterpriseArchitect` 实例，并设置 `project_root`。
#     *   `asyncTearDown` 方法:
#         *   清理临时测试目录。
#     *   `test_aico_p0_flow` 方法:
#         *   使用 `@patch('metagpt.actions.action.Action.run')` 装饰器来模拟所有 `Action` 的 `run` 方法。
#         *   `mock_llm_responses` 方法根据不同的 `Action` 类名，返回不同的模拟 LLM 响应。
#         *   按照测试用例的步骤，依次调用 `AICOProjectManager` 的方法，并使用 `self.assertTrue`、`self.assertEqual` 等断言方法来验证预期结果。
#         *   对生成的文档路径进行断言，确保文档已生成。 (对文档内容的详细断言可以进一步补充)
#         *   对跟踪表 (`project_tracking.xlsx`) 的操作进行了简化，只断言文件存在。 (可以根据需要添加更详细的断言，例如检查特定单元格的值)

# 2.  **`mock_llm_responses` 方法**:
#     *   根据传入的 `Action` 类名，返回不同的模拟 LLM 响应。
#     *   每个响应都是一个 `MagicMock` 对象，可以模拟 `ActionOutput` 的 `content` 和 `instruct_content` 属性。

# **运行测试**:

# 1.  确保已安装 `metagpt` 和 `pytest`。
# 2.  将上述代码保存为 `test_aico_p0.py` 文件。
# 3.  在命令行中，进入包含 `test_aico_p0.py` 文件的目录。
# 4.  运行 `python -m unittest test_aico_p0.py`。

# **注意事项**:

# *   这个测试用例代码只是一个基本框架，你需要根据实际情况补充更多的断言，以确保测试的完整性。
# *   `mock_llm_responses` 方法中的模拟响应只是示例，你需要根据你的 LLM 或 Agent 的实际输出来调整。
# *   对 `project_tracking.xlsx` 文件的断言可以进一步细化，例如使用 `openpyxl` 库来读取和验证 Excel 文件中的数据。
# *   可以添加更多的测试用例，覆盖不同的场景和边界条件。
# *   由于测试环境的限制，可能无法完全模拟真实环境中的所有情况，因此在实际部署前，还需要进行更全面的测试。
# *   确保你的 `metagpt` 版本是最新的, 以避免版本不兼容的问题。

# 通过这个测试用例代码，你可以自动化地测试 AICO 框架 P0 阶段的基本功能，从而提高开发效率和代码质量。