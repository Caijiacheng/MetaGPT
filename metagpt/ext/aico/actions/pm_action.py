#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Time    : 2023/12/15
@Author  : Jiacheng Cai
@File    : pm_action.py
说明:
  PM角色的核心动作包括:
  1. PrepareProject: 初始化项目,生成项目基本信息
  2. ReviewAllRequirements: 调用AI引擎复核需求,形成需求基线
  3. PlanSprintReleases: 调用AI引擎制定迭代计划
  4. ReviewAllDesigns: 调用AI引擎复核设计文档
  5. ReviewAllTasks: 调用AI引擎复核任务分解
"""
from pathlib import Path
from typing import Dict, List
from datetime import datetime
from openpyxl import Workbook, load_workbook
import logging
import shutil
from metagpt.actions import Action
import json
from metagpt.ext.aico.services.spec_service import SpecService

REVIEW_REQUIREMENTS_PROMPT = """作为资深项目经理，请根据以下材料进行需求复核：

【原始需求】
{raw_requirement}

【已解析需求】
{parsed_requirements}

【关联文档】
用户故事: {user_stories}
业务架构: {business_arch}
技术架构: {tech_arch}

请检查：
1. 存在的需求类型（业务/技术）是否解析完整
2. 需求与架构设计是否一致
3. 用户故事是否可追溯至需求（如存在业务需求）
4. 是否存在矛盾或缺失

输出格式：
{{
    "approved": true/false,
    "missing_items": ["缺失项列表"],
    "conflicts": ["矛盾点列表"],
    "trace_issues": ["可追溯性问题"],
    "suggestions": "改进建议"
}}"""

PLAN_SPRINT_PROMPT = """
你是一位项目经理,请调用AI引擎基于以下需求制定迭代计划,输出JSON格式:
{
    "sprint_plan": {
        "total_sprints": "迭代总数",
        "sprint_duration": "每个迭代的周期(周)",
        "sprints": [
            {
                "sprint_id": "迭代编号",
                "start_date": "开始日期",
                "end_date": "结束日期",
                "goals": ["目标1", "目标2"],
                "user_stories": ["故事1", "故事2"],
                "estimated_story_points": "预估故事点",
                "key_results": ["关键结果1", "关键结果2"]
            }
        ]
    }
}

需求和用户故事:
{requirements_and_stories}
"""

REVIEW_DESIGNS_PROMPT = """
你是一位项目经理,请调用AI引擎对以下设计文档进行复核,输出JSON格式的复核结果:
{
    "design_review": {
        "architecture_review": {
            "completeness": "完整度评分",
            "consistency": "一致性评分",
            "issues": ["问题1", "问题2"],
            "suggestions": ["建议1", "建议2"]
        },
        "interface_review": {
            "clarity": "清晰度评分",
            "standards": "规范性评分",
            "issues": ["问题1", "问题2"]
        },
        "data_model_review": {
            "rationality": "合理性评分",
            "issues": ["问题1", "问题2"]
        },
        "overall_assessment": "总体评估",
        "review_time": "YYYY-MM-DD HH:mm:ss"
    }
}

设计文档:
{designs}
"""

REVIEW_TASKS_PROMPT = """
你是一位项目经理,请调用AI引擎对以下任务分解进行复核,输出JSON格式的复核结果:
{
    "task_review": {
        "completeness": "是否覆盖所有需求",
        "granularity": "任务粒度是否合适",
        "dependencies": "任务依赖关系是否合理",
        "workload": "工作量是否合理",
        "issues": ["问题1", "问题2"],
        "suggestions": ["建议1", "建议2"],
        "review_time": "YYYY-MM-DD HH:mm:ss"
    }
}

任务列表:
{tasks}
"""

class PrepareProject(Action):
    """初始化项目"""
    def __init__(self):
        super().__init__()
        # 从规范文档加载表格结构定义
        self.REQUIRED_SHEETS = {
            "需求跟踪": {
                "原始需求": [  # 新增原始需求跟踪表
                    "需求文件", "需求类型", "添加时间", "当前状态",
                    "BA解析时间", "EA解析时间", "完成时间", "备注"
                ],
                "需求管理": [
                    "需求ID", "需求名称", "需求描述", "需求来源", 
                    "需求优先级", "需求状态", "提出人", "提出时间",
                    "目标完成时间", "验收标准", "备注"
                ],
                "用户故事管理": [
                    "用户故事ID", "关联需求ID", "用户故事名称", "用户故事描述",
                    "优先级", "状态", "验收标准", "创建时间", "备注"
                ]
            },
            "任务跟踪": {
                "任务跟踪": [
                    "任务ID", "关联需求ID", "关联用户故事ID", "任务名称", 
                    "任务描述", "任务类型", "负责人", "任务状态",
                    "计划开始时间", "计划结束时间", "实际开始时间", 
                    "实际结束时间", "备注"
                ]
            }
        }
    
    def _init_project_dirs(self, project_root: Path) -> Dict[str, Path]:
        """初始化项目目录结构"""
        # 创建主目录
        project_root.mkdir(parents=True, exist_ok=True)
        
        # 创建标准目录结构
        dirs = {
            "raw_requirements": project_root / "raw_requirements",  # 原始需求目录
            "docs": {
                "root": project_root / "docs",  # 文档根目录
                "specs": project_root / "docs" / "specs",  # 规范文档目录
                "requirements": project_root / "docs" / "requirements",  # 需求文档目录
                "designs": project_root / "docs" / "designs",  # 设计文档目录
                "reports": project_root / "docs" / "reports"  # 报告文档目录
            },
            "tracking": project_root / "tracking",  # 跟踪文件目录
            "output": project_root / "output"  # 输出目录
        }
        
        # 创建目录
        for dir_path in [dirs["raw_requirements"], dirs["tracking"], dirs["output"]]:
            dir_path.mkdir(exist_ok=True)
            
        # 创建文档目录结构
        for doc_dir in dirs["docs"].values():
            doc_dir.mkdir(parents=True, exist_ok=True)
            
        return dirs
    
    def _copy_spec_files(self, spec_dir: Path):
        """复制规范文档到项目目录"""
        # 复制需求和任务跟踪规范
        src_req_spec = Path("docs/aico/norm/Req&Task-Tracking.md")
        if src_req_spec.exists():
            shutil.copy2(src_req_spec, spec_dir / "Req&Task-Tracking.md")
            
        # 复制其他规范文档
        # ...
    
    def _save_raw_requirement(self, 
                            raw_req: str, 
                            raw_req_dir: Path,
                            is_file: bool = False,
                            iteration: int = 1) -> Path:
        """保存原始需求
        
        Args:
            raw_req: 原始需求内容或文件路径
            raw_req_dir: 原始需求目录
            is_file: 是否是文件
            iteration: 迭代次数
        """
        if is_file:
            # 如果是文件路径,复制到原始需求目录
            src_path = Path(raw_req)
            # 添加迭代号到文件名
            dst_path = raw_req_dir / f"iter{iteration}_{src_path.name}"
            shutil.copy2(src_path, dst_path)
        else:
            # 如果是文本,创建新文件保存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            dst_path = raw_req_dir / f"iter{iteration}_requirement_{timestamp}.txt"
            with open(dst_path, "w", encoding="utf-8") as f:
                f.write(raw_req)
                
        return dst_path
    
    def _create_or_update_tracking_files(self, tracking_dir: Path) -> Dict[str, Path]:
        """创建或更新跟踪文件"""
        tracking_files = {}
        
        # 处理需求跟踪表
        req_tracking = tracking_dir / "ReqTracking.xlsx"
        if req_tracking.exists():
            # 如果文件存在,加载现有文件
            wb_req = load_workbook(req_tracking)
            # 检查并添加缺失的sheet
            for sheet_name, headers in self.REQUIRED_SHEETS["需求跟踪"].items():
                if sheet_name not in wb_req.sheetnames:
                    ws = wb_req.create_sheet(sheet_name)
                    ws.append(headers)
        else:
            # 创建新文件
            wb_req = Workbook()
            for sheet_name, headers in self.REQUIRED_SHEETS["需求跟踪"].items():
                if sheet_name == "需求管理":
                    ws = wb_req.active
                    ws.title = sheet_name
                else:
                    ws = wb_req.create_sheet(sheet_name)
                ws.append(headers)
        wb_req.save(req_tracking)
        tracking_files["req_tracking"] = req_tracking
        
        # 处理任务跟踪表
        task_tracking = tracking_dir / "TaskTracking.xlsx"
        if task_tracking.exists():
            # 如果文件存在,加载现有文件
            wb_task = load_workbook(task_tracking)
            # 检查并添加缺失的sheet
            if "任务跟踪" not in wb_task.sheetnames:
                ws = wb_task.create_sheet("任务跟踪")
                ws.append(self.REQUIRED_SHEETS["任务跟踪"]["任务跟踪"])
        else:
            # 创建新文件
            wb_task = Workbook()
            ws = wb_task.active
            ws.title = "任务跟踪"
            ws.append(self.REQUIRED_SHEETS["任务跟踪"]["任务跟踪"])
        wb_task.save(task_tracking)
        tracking_files["task_tracking"] = task_tracking
        
        return tracking_files
        
    async def run(self, project_info: Dict) -> Dict:
        """初始化或更新项目
        
        Args:
            project_info: 项目信息
                - name: 项目名称
                - output_root: 输出根目录
                - raw_requirement: 原始需求(文件路径或文本)
                - iteration: 当前迭代次数(默认为1)
        """
        # 1. 确定项目根目录
        project_root = Path(project_info["output_root"]) / project_info["name"]
        
        # 2. 初始化/确认目录结构
        dirs = self._init_project_dirs(project_root)
        
        # 3. 复制规范文档(首次初始化时)
        if not (dirs["docs"]["specs"] / "Req&Task-Tracking.md").exists():
            self._copy_spec_files(dirs["docs"]["specs"])
        
        # 4. 保存原始需求(带迭代号)
        raw_req = project_info["raw_requirement"]
        iteration = project_info.get("iteration", 1)
        try:
            # 尝试作为文件路径处理
            Path(raw_req).resolve(strict=True)
            is_file = True
        except:
            is_file = False
        
        req_file = self._save_raw_requirement(
            raw_req,
            dirs["raw_requirements"],
            is_file,
            iteration
        )
        
        # 5. 创建或更新跟踪文件
        tracking_files = self._create_or_update_tracking_files(dirs["tracking"])
        
        # 6. 返回项目配置信息
        return {
            "project_name": project_info["name"],
            "project_root": str(project_root),
            "raw_requirement": str(req_file),
            "req_tracking": str(tracking_files["req_tracking"]),
            "task_tracking": str(tracking_files["task_tracking"]),
            "iteration": iteration,
            "directories": {
                "raw_requirements": str(dirs["raw_requirements"]),
                "docs": {k: str(v) for k, v in dirs["docs"].items()},
                "tracking": str(dirs["tracking"]),
                "output": str(dirs["output"])
            }
        }

    async def _run_impl(self, input_data: Dict) -> Dict:
        # 创建项目目录结构
        project_root = Path(input_data["project_name"])
        project_root.mkdir(exist_ok=True)
        
        # 创建项目规范模板
        INIT_TEMPLATE = """# 项目专属规范模板

## 接口规范
- 必须包含请求/响应示例
- 错误码需明确说明

## 数据规范
- 时间格式统一使用ISO 8601
- 金额单位统一为人民币分

## 其他要求
请在此补充项目特有规范...
"""
        
        project_spec_dir = project_root / "docs/specs"
        project_spec_dir.mkdir(parents=True, exist_ok=True)
        
        # 仅当不存在时创建
        for spec_type in ["ea_design", "project_tracking"]:
            spec_file = project_spec_dir / f"{spec_type}_spec.md"
            if not spec_file.exists():
                spec_file.write_text(INIT_TEMPLATE, encoding="utf-8")
        
        return {
            "project_root": str(project_root),
            "spec_dir": str(project_spec_dir)
        }

class ReviewAllRequirements(Action):
    """复核需求"""
    async def _run_impl(self, input_data: Dict) -> Dict:
        # 从输入数据中提取各类文档
        raw_req = input_data.get("raw_requirement", "")
        parsed_req = input_data.get("parsed_requirements", {})
        user_stories = input_data.get("user_stories", "")
        business_arch = input_data.get("business_arch", "")
        tech_arch = input_data.get("tech_arch", "")
        
        prompt = REVIEW_REQUIREMENTS_PROMPT.format(
            raw_requirement=raw_req,
            parsed_requirements=json.dumps(parsed_req, indent=2),
            user_stories=user_stories,
            business_arch=business_arch,
            tech_arch=tech_arch
        )
        
        result = await self.llm.aask(prompt)
        return self._parse_result(result)
        
    def _parse_result(self, raw: str) -> Dict:
        try:
            data = json.loads(raw)
            return {
                "approved": data.get("approved", False),
                "issues": {
                    "missing": data.get("missing_items", []),
                    "conflicts": data.get("conflicts", []),
                    "traceability": data.get("trace_issues", [])
                },
                "suggestions": data.get("suggestions", "")
            }
        except Exception as e:
            return {
                "approved": False,
                "issues": {"parse_error": [f"JSON解析失败: {str(e)}"]},
                "suggestions": "请检查AI输出格式"
            }

class PlanSprintReleases(Action):
    """制定迭代计划"""
    async def run(self, requirements_and_stories: Dict) -> Dict:
        # 调用AI引擎制定迭代计划
        prompt = PLAN_SPRINT_PROMPT.format(
            requirements_and_stories=requirements_and_stories
        )
        sprint_plan = await self.llm.aask(prompt)
        return sprint_plan

class ReviewAllDesigns(Action):
    """复核设计文档"""
    async def run(self, designs: Dict) -> Dict:
        # 调用AI引擎复核设计
        prompt = REVIEW_DESIGNS_PROMPT.format(
            designs=designs
        )
        review_result = await self.llm.aask(prompt)
        return review_result

class ReviewAllTasks(Action):
    """复核任务分解"""
    async def run(self, tasks: Dict) -> Dict:
        # 调用AI引擎复核任务
        prompt = REVIEW_TASKS_PROMPT.format(
            tasks=tasks
        )
        review_result = await self.llm.aask(prompt)
        return review_result 