from pathlib import Path
from openpyxl import load_workbook, Workbook
from datetime import datetime
from enum import Enum
import logging
from typing import Optional, Dict, List
from dataclasses import dataclass
from collections import defaultdict
import re
from functools import wraps
import threading

from metagpt.ext.aico.config.tracking_config import SheetType, IDConfig


logger = logging.getLogger(__name__)

class TrackingSheet(Enum):
    """跟踪表工作表定义"""
    RAW_REQUIREMENTS = "原始需求"
    REQUIREMENT_MGMT = "需求管理"
    USER_STORY_MGMT = "用户故事管理"
    TASK_TRACKING = "任务跟踪"

@dataclass
class TrackingServiceConfig:
    file_path: Path
    env: str = "prod"

def validate_id_format(pattern: str):
    """ID格式验证装饰器"""
    def decorator(func):
        @wraps(func)
        def wrapper(self, req_data: dict):
            req_id = req_data.get("id")
            if req_id and not re.match(pattern, req_id):
                raise ValueError(f"ID格式不合法: {req_id}")
            return func(self, req_data)
        return wrapper
    return decorator

class ProjectTrackingManager:
    """项目跟踪服务（Excel实现）"""
    
    @classmethod
    def from_path(cls, path: Path, **kwargs):
        """替代构造函数：从路径初始化"""
        return cls(path, **kwargs)
    def __init__(self, file_path: Path):
        self.file_path = Path(file_path) if isinstance(file_path, str) else file_path

        self.wb = self._init_workbook()
        self._validate_column_index(SheetType.RAW_REQUIREMENT.value.columns)
        self._validate_column_index(SheetType.REQUIREMENT_MGMT.value.columns)
        self._setup_sheets()
        self.requirement_counter = self._load_counter("req")
        self.user_story_counter = self._load_user_story_counters()
        self.task_counter = self._load_task_counters()
        self._lock = threading.Lock()
        
    def _init_workbook(self) -> Workbook:
        """初始化工作簿（修复文件保存问题）"""
        if self.file_path.exists():
            return load_workbook(self.file_path)
        else:
            wb = Workbook()
            wb.save(self.file_path)  # 立即保存新文件
            return wb
    
    def _setup_sheets(self):
        """根据配置初始化工作表（修复required_sheets定义）"""
        required_sheets = {
            st.value.name: st.value.headers 
            for st in SheetType 
            # 确保SheetType枚举中已移除VERSION_HISTORY
        }
        
        # 创建缺失的工作表
        for sheet_name, headers in required_sheets.items():
            if sheet_name not in self.wb.sheetnames:
                ws = self.wb.create_sheet(sheet_name)
                ws.append(headers)
        
        # 删除默认的空白表
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
            

    
    # 原始需求表操作 --------------------------------------------------
    def get_raw_requirement_status(self, file_path: str) -> Dict:
        """获取原始需求状态（补充complete_time）"""
        ws = self.wb[SheetType.RAW_REQUIREMENT.value.name]
        cols = SheetType.RAW_REQUIREMENT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if row[cols.RAW_FILE_PATH.value-1].value == file_path:
                return {
                    "status": row[cols.STATUS.value-1].value,
                    "ba_time": row[cols.BA_PARSE_TIME.value-1].value,
                    "complete_time": row[cols.COMPLETE_TIME.value-1].value  # 新增字段
                }
        return {"status": "待分析", "ba_time": None, "complete_time": None}
        
    _VALID_TRANSITIONS = {
        "待分析": ["已分析", "已驳回"],
        "已分析": ["完成", "已驳回"],
        "已驳回": ["已分析"],
        "完成": []
    }

    def update_raw_requirement_status(self, file_path: str, new_status: str):
        """更新需求状态（完整实现）"""
        current_status = self.get_raw_requirement_status(file_path)["status"]
        
        # 状态流转验证
        if new_status not in self._VALID_TRANSITIONS.get(current_status, []):
            raise ValueError(f"非法状态流转: {current_status} → {new_status}")

        # 完整更新逻辑
        ws = self.wb[SheetType.RAW_REQUIREMENT.value.name]
        cols = SheetType.RAW_REQUIREMENT.value.columns
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for row in ws.iter_rows(min_row=2):
            if row[cols.RAW_FILE_PATH.value-1].value == file_path:
                # 更新状态字段
                row[cols.STATUS.value-1].value = new_status
                
                # 记录时间戳
                if new_status == "已分析":
                    row[cols.BA_PARSE_TIME.value-1].value = now
                elif new_status == "完成":
                    row[cols.COMPLETE_TIME.value-1].value = now
                
                self.wb.save(self.file_path)
                return True
        return False
    
    # 需求管理表操作 --------------------------------------------------
    @validate_id_format(r"^REQ-PM\d{6}$")
    def add_standard_requirement(self, data: dict):
        """添加需求时处理优先级字段（增加name校验）"""
        if "name" not in data:
            raise ValueError("需求名称不能为空")
        
        data.setdefault("priority", IDConfig.DEFAULT_PRIORITY)
        
        if data["priority"] not in ["高", "中", "低"]:
            raise ValueError("优先级必须为高/中/低")
        
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        # 生成需求数据
        req_data = {
            "id": self._generate_req_id(),
            "raw_file_ref": data.get("raw_file", ""),
            "name": data["name"],
            "description": data.get("description", ""),
            "source": data.get("source", "内部需求"),
            "priority": data["priority"],
            "status": data.get("status", "已提出"),
            "owner": data.get("owner", ""),
            "submit_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "due_date": data.get("due_date", ""),
            "acceptance": data.get("acceptance", ""),
            "notes": data.get("notes", "")
        }
        
        # 写入Excel
        new_row = [req_data[key] for key in [
            "id", "raw_file_ref", "name", "description", "source",
            "priority", "status", "owner", "submit_time", "due_date",
            "acceptance", "notes"
        ]]
        ws.append(new_row)
        self.wb.save(self.file_path)
        
        return req_data  # 返回完整的需求数据
    
    # 用户故事管理操作 --------------------------------------------------
    def add_user_story(self, story_data: dict):
        """添加用户故事记录"""
        # 生成用户故事ID
        req_id = story_data.get("related_req_id")
        if not req_id:
            raise ValueError("必须指定关联需求ID")
        story_id = self._generate_user_story_id(req_id)
        story_data["story_id"] = story_id
        
        # 设置默认状态
        story_data.setdefault("status", "待拆分")
        
        ws = self.wb[SheetType.USER_STORY.value.name]
        cols = SheetType.USER_STORY.value.columns
        
        new_row = [
            story_id,
            req_id,
            story_data.get("name"),
            story_data.get("description"),
            story_data.get("priority"),
            story_data.get("status"),
            story_data.get("acceptance"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            story_data.get("notes")
        ]
        ws.append(new_row)
        self.wb.save(self.file_path)
        return {  # 返回完整用户故事数据
            "story_id": story_id,
            "name": story_data.get("name"),
            "status": story_data.get("status")
        }

    def update_user_story_status(self, story_id: str, new_status: str):
        """更新用户故事状态（修复列索引）"""
        ws = self.wb[SheetType.USER_STORY.value.name]
        cols = SheetType.USER_STORY.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.STORY_ID.value-1].value) == story_id:  # 使用正确的列索引
                row[cols.STATUS.value-1].value = new_status
                self.wb.save(self.file_path)
                return True
        return False

    def add_task(self, task_data: dict):
        """添加任务记录"""
        # 生成任务ID
        story_id = task_data.get("related_story_id")
        task_type = task_data.get("type")
        if not story_id or not task_type:
            raise ValueError("必须指定关联用户故事ID和任务类型")
        
        task_id = self._generate_task_id(story_id, task_type)
        task_data["task_id"] = task_id
        
        # 设置默认状态
        task_data.setdefault("status", "待开始")
        
        ws = self.wb[SheetType.TASK_TRACKING.value.name]
        cols = SheetType.TASK_TRACKING.value.columns
        
        new_row = [
            task_id,
            task_data.get("related_req_id"),
            story_id,
            task_data.get("name"),
            task_data.get("description"),
            task_type,
            task_data.get("owner"),
            task_data.get("status"),
            task_data.get("plan_start"),
            task_data.get("plan_end"),
            None,  # 实际开始时间
            None,  # 实际结束时间
            task_data.get("artifacts"),
            task_data.get("notes")
        ]
        ws.append(new_row)
        self.wb.save(self.file_path)
        return {  # 返回完整任务数据
            "task_id": task_id,
            "status": task_data.get("status"),
            "owner": task_data.get("owner")
        }

    def update_task_status(self, task_id: str, status: str, update_time: datetime = None):
        """更新任务状态"""
        ws = self.wb[SheetType.TASK_TRACKING.value.name]
        cols = SheetType.TASK_TRACKING.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.TASK_ID.value-1].value) == task_id:  # 修正索引
                row[cols.STATUS.value-1].value = status          # 修正索引
                if status == "进行中" and not row[cols.ACTUAL_START.value-1].value:  # 修正索引
                    row[cols.ACTUAL_START.value-1].value = update_time
                elif status == "已完成":
                    row[cols.ACTUAL_END.value-1].value = update_time
                self.wb.save(self.file_path)
                return True
        return False

    def add_raw_requirement(
        self,
        file_path: str,
        req_type: str,
        status: str = "待分析",
        comment: str = ""
    ):
        """添加原始需求记录"""
        if not file_path.startswith("raw/iter"):
            raise ValueError("原始需求文件路径必须符合raw/iterX/格式")
        ws = self.wb[SheetType.RAW_REQUIREMENT.value.name]
        cols = SheetType.RAW_REQUIREMENT.value.columns
        
        ws.append([
            file_path,
            req_type,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            status,
            "",  # 关联需求ID
            None,  # BA解析时间
            None,  # EA解析时间
            None,  # 完成时间
            comment
        ])
        self.wb.save(self.file_path)
        
    def get_pending_requirements(self) -> List[dict]:
        """获取待分析需求"""
        ws = self.wb[SheetType.RAW_REQUIREMENT.value.name]
        cols = SheetType.RAW_REQUIREMENT.value.columns
        
        pending_reqs = []
        for row in ws.iter_rows(min_row=2):
            if row[cols.STATUS.value-1].value == "待分析":  # 修正索引
                pending_reqs.append({
                    "file_path": row[cols.RAW_FILE_PATH.value-1].value,  # 修正索引
                    "req_type": row[cols.REQ_TYPE.value-1].value,        # 修正索引
                    "add_time": row[cols.ADD_TIME.value-1].value         # 修正索引
                })
        return pending_reqs
    
    def update_requirement_status(self, req_id: str, status: str):
        """更新需求状态"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.REQ_ID.value-1].value) == req_id:
                row[cols.REQ_STATUS.value-1].value = status
                self.wb.save(self.file_path)
                return True
        return False

    def validate_file_structure(self) -> bool:
        """校验文件结构完整性"""
        required_sheets = {st.value.name for st in SheetType}
        existing_sheets = set(self.wb.sheetnames)
        
        # 检查必需工作表是否存在
        if not required_sheets.issubset(existing_sheets):
            return False
        
        # 检查各表头是否正确
        for sheet_type in SheetType:
            ws = self.wb[sheet_type.value.name]
            expected_headers = sheet_type.value.headers
            actual_headers = [cell.value for cell in ws[1]]  # 第一行为表头
            if actual_headers != expected_headers:
                return False
        
        return True

    def update_requirement(self, req_id: str, updates: dict) -> dict:
        """更新需求信息（修复字段映射）"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        # 字段映射修正
        column_mapping = {
            "priority": cols.REQ_PRIORITY,
            "status": cols.REQ_STATUS
        }

        for row in ws.iter_rows(min_row=2):
            if str(row[cols.REQ_ID.value-1].value) == req_id:
                for field, value in updates.items():
                    if field in column_mapping:  # 添加字段存在性检查
                        col = column_mapping[field]
                        row[col.value-1].value = value
                self.wb.save(self.file_path)
                return self.get_requirement(req_id)
        return None

    def _get_column_index(self, sheet_type: SheetType, header: str) -> int:
        """动态获取列索引"""
        config = sheet_type.value
        try:
            return config.headers.index(header) + 1  # Excel列从1开始
        except ValueError:
            raise KeyError(f"列'{header}'不存在于工作表{config.name}")

    def validate_raw_requirement(self, req_data: dict) -> bool:
        """校验原始需求记录完整性（文档5.2节）"""
        required_fields = ["file_path", "description", "source"]
        return all(field in req_data for field in required_fields)

    def update_design_status(self, req_id: str, doc_type: str, doc_path: str, version: str):
        """更新设计状态（根据反馈调整参数）"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        # 根据文档类型更新不同字段
        for row in ws.iter_rows(min_row=2):
            if row[cols.REQ_ID.value-1].value == req_id:
                if doc_type == "PRD":
                    row[cols.PRD_STATUS.value-1].value = "基线化"
                    row[cols.PRD_DOC.value-1].value = doc_path
                elif doc_type == "TECH_DESIGN":
                    row[cols.TECH_DESIGN_STATUS.value-1].value = "基线化"
                    row[cols.TECH_DOC.value-1].value = doc_path
                
                row[cols.DESIGN_VERSION.value-1].value = version
                self.wb.save(self.file_path)
                return True
        return False

    def update_task_artifacts(self, task_id: str, artifacts: list):
        """更新任务关联产出物"""
        ws = self.wb[SheetType.TASK_TRACKING.value.name]
        cols = SheetType.TASK_TRACKING.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.TASK_ID.value-1].value) == task_id:
                existing = row[cols.ARTIFACTS.value-1].value or ""
                new_artifacts = existing.split(";") + artifacts
                row[cols.ARTIFACTS.value-1].value = ";".join(filter(None, new_artifacts))
                self.wb.save(self.file_path)
                return True
        return False

    def mark_baseline(self, baseline_type: str, version: str, req_ids: list):
        """标记需求基线"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            req_id = str(row[cols.REQ_ID.value-1].value)  # 保持已修正的索引
            if req_id in req_ids:
                if baseline_type == "design":
                    row[cols.DESIGN_STATUS.value-1].value = "基线化"          # 修正索引
                    row[cols.DESIGN_REVIEW.value-1].value = datetime.now().isoformat()  # 修正索引
                elif baseline_type == "code":
                    row[cols.CODE_BASELINE.value-1].value = version         # 修正索引
        self.wb.save(self.file_path)
        
    def get_baseline_requirements(self, baseline_type: str, version: str = None) -> list:
        """获取基线需求"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        baseline_reqs = []
        for row in ws.iter_rows(min_row=2):
            if baseline_type == "design" and row[cols.DESIGN_STATUS.value-1].value == "基线化":
                if not version or row[cols.DESIGN_REVIEW.value-1].value.startswith(version):
                    baseline_reqs.append(row[cols.REQ_ID.value-1].value)
            elif baseline_type == "code" and row[cols.CODE_BASELINE.value-1].value == version:
                baseline_reqs.append(row[cols.REQ_ID.value-1].value)
        return baseline_reqs

    def _validate_column_index(self, cols: Enum):
        """校验列索引有效性"""
        for col in cols:
            if col.value < 1 or col.value > 20:  # 假设最大列数20
                raise ValueError(f"无效列索引 {col.name}={col.value}")
    
    def _generate_req_id(self) -> str:
        """生成需求ID（完整线程安全实现）"""
        with self._lock:
            now = datetime.now()
            seq = self.requirement_counter
            self.requirement_counter += 1
            return f"REQ-PM{now:%y%m}{seq:03d}"
    
    def _generate_user_story_id(self, req_id: str) -> str:
        """生成用户故事ID"""
        key = f"us_{req_id}"
        self.user_story_counter[key] = self.user_story_counter.get(key, 0) + 1
        return f"US-{req_id}-{self.user_story_counter[key]:02d}"
    
    def _generate_task_id(self, story_id: str, task_type: str) -> str:
        """生成任务ID"""
        key = f"task_{story_id}_{task_type}"
        self.task_counter[key] = self.task_counter.get(key, 0) + 1
        return f"T-{story_id}-{IDConfig.TASK_TYPES[task_type]}{self.task_counter[key]:02d}"
    
    def _load_counter(self, counter_type: str) -> int:
        """加载序列号（完整实现）"""
        if counter_type == "req":
            max_seq = 0
            pattern = re.compile(r"REQ-PM\d{4}(\d{3})$")  # 匹配7位日期+3位序列号
            for row in self.wb[SheetType.REQUIREMENT_MGMT.value.name].iter_rows(min_row=2):
                if (req_id := str(row[0].value)) and (match := pattern.match(req_id)):
                    seq = int(match.group(1))
                    max_seq = max(max_seq, seq)
            return max_seq + 1 if max_seq > 0 else 1
        return 1
    
    def _load_user_story_counters(self) -> defaultdict:
        """加载用户故事序列号"""
        counters = defaultdict(int)
        pattern = re.compile(rf"US-(REQ-{IDConfig.PROJECT_PREFIX}\d{{6}})-(\d{{2}})$")
        for row in self.wb[SheetType.USER_STORY.value.name].iter_rows(min_row=2):
            if match := pattern.match(str(row[0].value)):
                req_id = match.group(1)
                seq = int(match.group(2))
                counters[req_id] = max(counters[req_id], seq)
        return counters
    
    def _load_task_counters(self) -> defaultdict:
        """加载任务序列号"""
        counters = defaultdict(int)
        pattern = re.compile(rf"T-(US-REQ-{IDConfig.PROJECT_PREFIX}\d{{6}}-\d{{2}})-[A-Z]{{2,6}}(\d{{{IDConfig.TASK_SEQ_LENGTH}}})$")
        for row in self.wb[SheetType.TASK_TRACKING.value.name].iter_rows(min_row=2):
            if match := pattern.match(str(row[0].value)):
                us_id = match.group(1)
                seq = int(match.group(2))
                counters[us_id] = max(counters[us_id], seq)
        return counters
    

    def get_user_story(self, story_id: str) -> Optional[Dict]:
        """获取用户故事完整信息"""
        ws = self.wb[SheetType.USER_STORY.value.name]
        cols = SheetType.USER_STORY.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.STORY_ID.value-1].value) == story_id:
                return {
                    "story_id": story_id,
                    "related_req_id": row[cols.RELATED_REQ_ID.value-1].value,
                    "name": row[cols.STORY_NAME.value-1].value,
                    "description": row[cols.STORY_DESC.value-1].value,
                    "priority": row[cols.STORY_PRIORITY.value-1].value,
                    "status": row[cols.STATUS.value-1].value,
                    "create_time": row[cols.CREATE_TIME.value-1].value,
                    "notes": row[cols.STORY_NOTES.value-1].value
                }
        return None

    def get_requirement(self, req_id: str) -> Optional[Dict]:
        """获取需求完整信息（新增方法）"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.REQ_ID.value-1].value) == req_id:
                return {
                    "id": req_id,
                    "name": row[cols.REQ_NAME.value-1].value,
                    "priority": row[cols.REQ_PRIORITY.value-1].value,
                    "status": row[cols.REQ_STATUS.value-1].value,
                    # 其他字段...
                }
        return None

