from pathlib import Path
from openpyxl import load_workbook, Workbook
from datetime import datetime
from enum import Enum
import logging
from typing import Optional, Dict, List
from dataclasses import dataclass
from collections import defaultdict

from metagpt.ext.aico.config.tracking_config import SheetType
from metagpt.ext.aico.services.version_service import VersionService, get_current_version


logger = logging.getLogger(__name__)

class TrackingSheet(Enum):
    """跟踪表工作表定义"""
    RAW_REQUIREMENTS = "原始需求"
    REQUIREMENT_MGMT = "需求管理"
    USER_STORY_MGMT = "用户故事管理"
    TASK_TRACKING = "任务跟踪"

@dataclass
class TrackingServiceConfig:
    file_path: Path
    env: str = "prod"

class ProjectTrackingService:
    """项目跟踪服务（Excel实现）"""
    
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.wb = self._init_workbook()
        self._setup_sheets()
        
    def _init_workbook(self) -> Workbook:
        """初始化工作簿"""
        if self.file_path.exists():
            return load_workbook(self.file_path)
        return Workbook()
    
    def _setup_sheets(self):
        """根据配置初始化工作表"""
        required_sheets = {st.value.name: st.value.headers for st in SheetType}
        
        # 创建缺失的工作表
        for sheet_name, headers in required_sheets.items():
            if sheet_name not in self.wb.sheetnames:
                ws = self.wb.create_sheet(sheet_name)
                ws.append(headers)
        
        # 删除默认的空白表
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
            
    def add_version_record(
        self, 
        version: str,
        doc_path: str,
        doc_type: str,
        change_type: str,
        changes: List[str],
        related_reqs: List[str]
    ):
        """添加版本记录（增强版）"""
        current_version = get_current_version(self.project_root)
        if version != current_version:
            logger.error(f"版本记录{version}与VERSION文件{current_version}不一致")
            raise ValueError("版本记录不一致")
        
        ws = self.wb["版本历史"]
        ws.append([
            version,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            doc_path,
            doc_type,
            change_type,
            "\n".join(changes),
            ",".join(related_reqs)
        ])
        
    def get_latest_version(self, doc_type: str) -> str:
        """改为从VERSION文件获取"""
        return get_current_version(self.project_root)
    
    # 原始需求表操作 --------------------------------------------------
    def get_raw_requirement_status(self, file_path: str) -> Optional[Dict]:
        """获取原始需求状态"""
        ws = self.wb[SheetType.RAW_REQUIREMENT.value.name]
        cols = SheetType.RAW_REQUIREMENT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if row[cols.RAW_FILE_PATH.value].value == file_path:
                return {
                    "status": row[cols.STATUS.value].value,
                    "ba_time": row[cols.BA_PARSE_TIME.value].value
                }
        return None
        
    def update_raw_requirement_status(
        self, 
        file_path: str, 
        status: str,
        update_time: datetime = datetime.now()
    ):
        """更新原始需求状态"""
        ws = self.wb[SheetType.RAW_REQUIREMENT.value.name]
        cols = SheetType.RAW_REQUIREMENT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if row[cols.RAW_FILE_PATH.value].value == file_path:
                row[cols.STATUS.value].value = status
                if status == "已分析":
                    row[cols.BA_PARSE_TIME.value].value = update_time
                self.wb.save(self.file_path)
                return True
        return False
    
    # 需求管理表操作 --------------------------------------------------
    def add_standard_requirement(self, req_data: dict):
        """添加标准需求记录"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        new_row = [
            req_data.get("id"),
            req_data.get("raw_file"),
            req_data.get("name"),
            req_data.get("description"),
            req_data.get("source"),
            req_data.get("priority"),
            req_data.get("status"),
            req_data.get("owner"),
            req_data.get("submit_time"),
            req_data.get("due_date"),
            req_data.get("acceptance"),
            req_data.get("notes")
        ]
        ws.append(new_row)
    
    # 用户故事管理操作 --------------------------------------------------
    def add_user_story(self, story_data: dict):
        """添加用户故事记录"""
        ws = self.wb[SheetType.USER_STORY.value.name]
        cols = SheetType.USER_STORY.value.columns
        
        new_row = [
            story_data.get("story_id"),
            story_data.get("related_req_id"),
            story_data.get("name"),
            story_data.get("description"),
            story_data.get("priority"),
            story_data.get("status"),
            story_data.get("acceptance"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            story_data.get("notes")
        ]
        ws.append(new_row)

    def update_user_story_status(self, story_id: str, status: str):
        """更新用户故事状态"""
        ws = self.wb[SheetType.USER_STORY.value.name]
        cols = SheetType.USER_STORY.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.STORY_ID.value].value) == story_id:
                row[cols.STATUS.value].value = status
                self.wb.save(self.file_path)
                return True
        return False

    def add_task(self, task_data: dict):
        """添加任务记录"""
        ws = self.wb[SheetType.TASK_TRACKING.value.name]
        cols = SheetType.TASK_TRACKING.value.columns
        
        new_row = [
            task_data.get("task_id"),
            task_data.get("related_req_id"),
            task_data.get("related_story_id"),
            task_data.get("name"),
            task_data.get("description"),
            task_data.get("type"),
            task_data.get("owner"),
            "待开始",  # 默认状态
            task_data.get("plan_start"),
            task_data.get("plan_end"),
            None,  # 实际开始时间
            None,  # 实际结束时间
            task_data.get("notes")
        ]
        ws.append(new_row)
        self.wb.save(self.file_path)

    def update_task_status(self, task_id: str, status: str, update_time: datetime = datetime.now()):
        """更新任务状态"""
        ws = self.wb[SheetType.TASK_TRACKING.value.name]
        cols = SheetType.TASK_TRACKING.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.TASK_ID.value].value) == task_id:
                row[cols.TASK_STATUS.value].value = status
                if status == "进行中" and not row[cols.ACTUAL_START.value].value:
                    row[cols.ACTUAL_START.value].value = update_time
                elif status == "已完成":
                    row[cols.ACTUAL_END.value].value = update_time
                self.wb.save(self.file_path)
                return True
        return False

    def add_raw_requirement(
        self,
        file_path: str,
        req_type: str,
        status: str = "待分析",
        comment: str = ""
    ):
        """添加原始需求记录"""
        ws = self.wb[SheetType.RAW_REQUIREMENT.value.name]
        cols = SheetType.RAW_REQUIREMENT.value.columns
        
        ws.append([
            file_path,
            req_type,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            status,
            "",  # 关联需求ID
            None,  # BA解析时间
            None,  # EA解析时间
            None,  # 完成时间
            comment
        ])
        self.wb.save(self.file_path)
        
    def get_pending_requirements(self) -> List[dict]:
        """获取待分析需求"""
        ws = self.wb[SheetType.RAW_REQUIREMENT.value.name]
        cols = SheetType.RAW_REQUIREMENT.value.columns
        
        pending_reqs = []
        for row in ws.iter_rows(min_row=2):
            if row[cols.STATUS.value].value == "待分析":
                pending_reqs.append({
                    "file_path": row[cols.RAW_FILE_PATH.value].value,
                    "req_type": row[cols.REQ_TYPE.value].value,
                    "add_time": row[cols.ADD_TIME.value].value
                })
        return pending_reqs
    
    def update_requirement_status(self, req_id: str, status: str):
        """更新需求状态"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.REQ_ID.value].value) == req_id:
                row[cols.REQ_STATUS.value].value = status
                self.wb.save(self.file_path)
                return True
        return False

    def validate_file_structure(self) -> bool:
        """校验文件结构完整性"""
        required_sheets = {st.value.name for st in SheetType}
        existing_sheets = set(self.wb.sheetnames)
        
        # 检查必需工作表是否存在
        if not required_sheets.issubset(existing_sheets):
            return False
        
        # 检查各表头是否正确
        for sheet_type in SheetType:
            ws = self.wb[sheet_type.value.name]
            expected_headers = sheet_type.value.headers
            actual_headers = [cell.value for cell in ws[1]]  # 第一行为表头
            if actual_headers != expected_headers:
                return False
        
        return True

    def update_requirement(self, req_id: str, **kwargs):
        """更新需求信息"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.REQ_ID.value].value) == req_id:
                for field, value in kwargs.items():
                    col = getattr(cols, field.upper(), None)
                    if col and col.name.lower() == field.lower():
                        row[col.value].value = value
                self.wb.save(self.file_path)
                return True
        return False

    def _get_column_index(self, sheet_type: SheetType, header: str) -> int:
        """动态获取列索引"""
        config = sheet_type.value
        try:
            return config.headers.index(header) + 1  # Excel列从1开始
        except ValueError:
            raise KeyError(f"列'{header}'不存在于工作表{config.name}")

    def add_project_artifact(self, artifact_type: str, version: str, file_path: str):
        """记录项目产物"""
        ws = self.wb[SheetType.VERSION_HISTORY.value.name]
        cols = SheetType.VERSION_HISTORY.value.columns
        
        ws.append([
            version,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            str(file_path),
            artifact_type,
            "产物生成",
            f"新增{artifact_type}产物: {version}",
            ""
        ])
        self.wb.save(self.file_path)

    def get_version_history(self, version_filter: dict = None) -> List[dict]:
        """获取带过滤条件的版本历史"""
        ws = self.wb[SheetType.VERSION_HISTORY.value.name]
        cols = SheetType.VERSION_HISTORY.value.columns
        
        history = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {
                "version": row[cols.VERSION.value-1],
                "time": row[cols.GEN_TIME.value-1],
                "doc_path": row[cols.DOC_PATH.value-1],
                "doc_type": row[cols.DOC_TYPE.value-1],
                "change_type": row[cols.CHANGE_TYPE.value-1],
                "changes": row[cols.CHANGES.value-1].split("\n") if row[cols.CHANGES.value-1] else [],
                "related_reqs": row[cols.RELATED_REQS.value-1].split(",") if row[cols.RELATED_REQS.value-1] else []
            }
            
            # 添加过滤逻辑
            if version_filter:
                match = all(
                    str(record.get(k, "")) == str(v) 
                    for k, v in version_filter.items()
                )
                if not match:
                    continue
                
            history.append(record)
        return sorted(history, key=lambda x: x["time"], reverse=True)

    def clean_old_versions(self, retention_policy: dict = None):
        """根据保留策略清理旧版本"""
        policy = retention_policy or {
            "major": 2,    # 保留最近2个主版本
            "minor": 3,     # 每个主版本保留3个次版本
            "patch": 2      # 每个次版本保留2个修订版
        }
        
        versions = self.get_version_history()
        version_map = defaultdict(lambda: defaultdict(list))
        
        # 构建版本映射
        for ver in versions:
            major, minor, patch = map(int, ver["version"].split('.'))
            version_map[major][minor].append(patch)
        
        # 实施清理策略
        for major in list(version_map.keys()):
            # 保留最近N个主版本
            if major > sorted(version_map.keys())[-policy["major"]]:
                del version_map[major]
                continue
            
            for minor in list(version_map[major].keys()):
                # 保留最近M个次版本
                if minor > sorted(version_map[major].keys())[-policy["minor"]]:
                    del version_map[major][minor]
                    continue
                
                # 保留最近K个修订版
                version_map[major][minor] = sorted(version_map[major][minor])[-policy["patch"]:]
        
        # 生成有效版本列表
        valid_versions = set()
        for major, minors in version_map.items():
            for minor, patches in minors.items():
                for patch in patches:
                    valid_versions.add(f"{major}.{minor}.{patch}")
        
        # 清理无效版本记录
        ws = self.wb[SheetType.VERSION_HISTORY.value.name]
        for row_idx in range(ws.max_row, 1, -1):
            version = ws.cell(row=row_idx, column=1).value
            if version not in valid_versions:
                ws.delete_rows(row_idx)

    def validate_raw_requirement(self, req_data: dict) -> bool:
        """校验原始需求记录完整性（文档5.2节）"""
        required_fields = ["file_path", "description", "source"]
        return all(field in req_data for field in required_fields)

    def add_change(self, change_data: dict) -> str:
        """添加变更记录"""
        ws = self.wb[SheetType.CHANGE_MGMT.value.name]
        cols = SheetType.CHANGE_MGMT.value.columns
        
        change_id = f"CHG-{ws.max_row:03d}"
        ws.append([
            change_id,
            change_data.get("related_req_id"),
            change_data.get("change_type"),
            change_data.get("description"),
            change_data.get("impact_analysis"),
            f"{change_data.get('submitter')}/{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "待审批",
            "未开始",
            change_data.get("baseline_impact"),
            "",
            None
        ])
        self.wb.save(self.file_path)
        return change_id

    def update_change_status(self, change_id: str, status: str, comment: str = ""):
        """更新变更状态"""
        ws = self.wb[SheetType.CHANGE_MGMT.value.name]
        cols = SheetType.CHANGE_MGMT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if row[cols.CHANGE_ID.value-1].value == change_id:
                if status == "approved":
                    row[cols.APPROVAL_STATUS.value-1].value = "已批准"
                elif status == "rejected":
                    row[cols.APPROVAL_STATUS.value-1].value = "已拒绝"
                    row[cols.IMPLEMENT_STATUS.value-1].value = "已关闭"
                    row[cols.CLOSE_TIME.value-1].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                row[cols.IMPACT_ANALYSIS.value-1].value = comment
                self.wb.save(self.file_path)
                return True
        return False

    def update_design_status(self, req_id: str, status: str, design_doc: str = None):
        """更新需求设计状态"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.REQ_ID.value-1].value) == req_id:
                row[cols.DESIGN_STATUS.value-1].value = status
                if design_doc:
                    row[cols.DESIGN_DOC.value-1].value = design_doc
                if status == "基线化":
                    row[cols.DESIGN_REVIEW.value-1].value = datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
                self.wb.save(self.file_path)
                return True
        return False

    def update_task_artifacts(self, task_id: str, artifacts: list):
        """更新任务关联产出物"""
        ws = self.wb[SheetType.TASK_TRACKING.value.name]
        cols = SheetType.TASK_TRACKING.value.columns
        
        for row in ws.iter_rows(min_row=2):
            if str(row[cols.TASK_ID.value-1].value) == task_id:
                existing = row[cols.ARTIFACTS.value-1].value or ""
                new_artifacts = existing.split(";") + artifacts
                row[cols.ARTIFACTS.value-1].value = ";".join(filter(None, new_artifacts))
                self.wb.save(self.file_path)
                return True
        return False

    def mark_baseline(self, baseline_type: str, version: str, req_ids: list):
        """标记需求基线"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        for row in ws.iter_rows(min_row=2):
            req_id = str(row[cols.REQ_ID.value-1].value)
            if req_id in req_ids:
                if baseline_type == "design":
                    row[cols.DESIGN_STATUS.value-1].value = "基线化"
                    row[cols.DESIGN_REVIEW.value-1].value = datetime.now().isoformat()
                elif baseline_type == "code":
                    row[cols.CODE_BASELINE.value-1].value = version
        self.wb.save(self.file_path)
        
    def get_baseline_requirements(self, baseline_type: str, version: str = None) -> list:
        """获取基线需求"""
        ws = self.wb[SheetType.REQUIREMENT_MGMT.value.name]
        cols = SheetType.REQUIREMENT_MGMT.value.columns
        
        baseline_reqs = []
        for row in ws.iter_rows(min_row=2):
            if baseline_type == "design" and row[cols.DESIGN_STATUS.value-1].value == "基线化":
                if not version or row[cols.DESIGN_REVIEW.value-1].value.startswith(version):
                    baseline_reqs.append(row[cols.REQ_ID.value-1].value)
            elif baseline_type == "code" and row[cols.CODE_BASELINE.value-1].value == version:
                baseline_reqs.append(row[cols.REQ_ID.value-1].value)
        return baseline_reqs

