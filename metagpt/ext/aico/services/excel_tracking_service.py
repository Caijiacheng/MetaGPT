from openpyxl import Workbook
from pathlib import Path
from metagpt.ext.aico.config.tracking_config import SheetType
from .base_tracking_service import ITrackingService

class ExcelTrackingService(ITrackingService):
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self._wb = None
        
    def _get_workbook(self) -> Workbook:
        # 保持原有实现，但使用SheetType配置
        if not self._wb:
            if self.file_path.exists():
                self._wb = load_workbook(self.file_path)
            else:
                self._create_new_file()
        return self._wb
    
    def _create_new_file(self):
        self._wb = Workbook()
        self._init_sheets()
        self.save()
        
    def _init_sheets(self):
        """使用配置类初始化工作表"""
        if 'Sheet' in self._wb.sheetnames:
            del self._wb['Sheet']
            
        for sheet_type in SheetType:
            config = sheet_type.value
            if config.name not in self._wb.sheetnames:
                ws = self._wb.create_sheet(config.name)
                ws.append(config.headers)
    
    # 其他方法实现改用SheetType配置... 