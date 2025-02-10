#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Time    : 2023/12/14
@Author  : Jiacheng Cai
@File    : enterprise_architect.py
@Modified By: Jiacheng Cai, 2023/12/15
说明:
  企业架构师(EA)角色职责:
  1. 接收并分析技术需求(从架构文档中提取)
  2. 调用AI引擎生成技术需求矩阵
  3. 调用AI引擎更新4A架构(应用架构、数据架构、技术架构)
  4. 发布分析结果到环境中
"""

from metagpt.roles import Role
from metagpt.environment.aico.aico_env import AICOEnvironment
from ..actions.ea_action import ParseTechRequirements, Update4ATech
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
from metagpt.ext.aico.services.doc_manager import DocManagerService, DocType
import logging

logger = logging.getLogger(__name__)

class AICOEnterpriseArchitect(Role):
    """遵循AICO规范的企业架构师角色"""
    
    name: str = "AICO_EA"
    profile: str = "Enterprise Architect"
    goal: str = "设计技术架构并维护架构资产"
    constraints: str = "严格遵循AICO技术架构规范"
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.doc_manager = DocManagerService(self.project_root)
        
    async def _act(self) -> None:
        """纯消息驱动模式"""
        async for msg in self.observe(AICOEnvironment.MSG_REQUIREMENT_TECH_ANALYSIS.name):
            try:
                # 通知PM分析开始
                await self.publish(AICOEnvironment.MSG_EA_ANALYSIS_STARTED.name, {
                    "req_id": msg.content["req_id"],
                    "start_time": datetime.now().isoformat()
                })
                
                result = await self._process_tech_requirement(msg.content)
                
                # 返回处理结果给PM
                await self.publish(AICOEnvironment.MSG_EA_ANALYSIS_DONE.name, {
                    "req_id": msg.content["req_id"],
                    "architecture": result["architecture"],
                    "output_files": [
                        str(result["architecture_file"]),
                        result["requirement_matrix_file"]
                    ],
                    "change_log": result["change_log"]
                })
                
            except Exception as e:
                logger.error(f"技术需求分析失败: {msg.content['req_id']}")
                await self.publish(AICOEnvironment.MSG_EA_ANALYSIS_FAILED.name, {
                    "req_id": msg.content["req_id"],
                    "error": str(e),
                    "fail_time": datetime.now().isoformat()
                })

    async def _process_tech_requirement(self, req_info: dict) -> dict:
        """处理技术需求分析核心逻辑"""
        # 执行需求解析
        parse_result = await self.rc.run(
            ParseTechRequirements().run({
                "raw_requirement": req_info["requirements"],
                "project_root": str(self.project_root),
                "req_id": req_info["req_id"],
                "version": req_info["version"]
            })
        )
        
        # 更新技术架构
        update_result = await self.rc.run(
            Update4ATech().run({
                "requirement_matrix": parse_result.instruct_content,
                "current_architecture": self._get_current_architecture(req_info["version"]),
                "project_root": str(self.project_root),
                "req_id": req_info["req_id"],
                "version": req_info["version"]
            })
        )
        
        return {
            "architecture": update_result.instruct_content["architecture"],
            "architecture_file": self._save_tech_architecture(
                req_info["version"],
                update_result.instruct_content
            ),
            "requirement_matrix_file": parse_result.content["output_file"],
            "change_log": update_result.instruct_content["change_log"]
        }

    def _get_current_architecture(self, version: str) -> str:
        """获取当前版本架构"""
        arch_path = self.doc_manager.get_doc_path(
            DocType.TECH_ARCH,
            version=version,
            create_dir=False
        )
        return arch_path.read_text(encoding="utf-8") if arch_path.exists() else ""

    def _save_tech_architecture(self, version: str, data: dict) -> Path:
        """保存技术架构文档"""
        return self.doc_manager.save_document(
            doc_type=DocType.TECH_ARCH,
            content=data["architecture"],
            version=version,
            metadata={
                "components": data["components"],
                "dependencies": data["dependencies"],
                "change_log": data["change_log"]
            }
        )

    def get_actions(self) -> list:
        return [
            ("parse_tech_requirements", ParseTechRequirements),
            ("update_4a_tech", Update4ATech)
        ]

    def _update_req_status(self, tracking_file: Path, status: str, req_id: str):
        """更新需求跟踪表状态"""
        wb = load_workbook(tracking_file)
        ws = wb["需求管理"]
        
        for row in ws.iter_rows(min_row=2):
            if row[0].value == req_id:  # 需求ID列
                row[3].value = status  # 状态列
                row[5].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # EA解析时间
                break
            
        wb.save(tracking_file)
